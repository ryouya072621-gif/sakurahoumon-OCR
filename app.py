import os
import json
import csv
import io
import time
import shutil
import zipfile
from pathlib import Path

from flask import (
    Flask, render_template, request, redirect, url_for,
    jsonify, send_file, abort
)

import ocr_engine
import ai_corrector

# Load .env file
from pathlib import Path as _Path
_env_file = _Path(__file__).parent / ".env"
if _env_file.exists():
    for line in _env_file.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = os.path.join(os.path.dirname(__file__), "uploads")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

ALLOWED_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "tif", "tiff"}


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route("/health")
def health():
    return jsonify({"status": "ok"})


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return redirect(url_for("index"))

    file = request.files["file"]
    if file.filename == "" or not allowed_file(file.filename):
        return redirect(url_for("index"))

    # Save with timestamp to avoid conflicts
    ext = file.filename.rsplit(".", 1)[1].lower()
    safe_name = f"upload_{int(time.time())}.{ext}"
    file_path = os.path.join(app.config["UPLOAD_FOLDER"], safe_name)
    file.save(file_path)

    # Create job and start background OCR
    job_id = ocr_engine.create_job(file_path)
    ocr_engine.start_ocr(job_id)

    return redirect(url_for("processing", job_id=job_id))


@app.route("/processing/<job_id>")
def processing(job_id):
    job = ocr_engine.get_job(job_id)
    if not job:
        abort(404)
    if job["status"] == "done":
        return redirect(url_for("results", job_id=job_id))
    return render_template("processing.html", job_id=job_id)


@app.route("/status/<job_id>")
def status(job_id):
    job = ocr_engine.get_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({
        "status": job["status"],
        "current_page": job["current_page"],
        "page_count": job["page_count"],
        "error": job["error"],
    })


@app.route("/results/<job_id>")
def results(job_id):
    job = ocr_engine.get_job(job_id)
    if not job:
        abort(404)
    if job["status"] != "done":
        return redirect(url_for("processing", job_id=job_id))

    return render_template(
        "results.html",
        job_id=job_id,
        page_count=len(job["results"]),
        ocr_data=job["results"],
    )


@app.route("/image/<job_id>/<int:page>")
def page_image(job_id, page):
    buf = ocr_engine.get_page_image_jpeg(job_id, page)
    if buf is None:
        abort(404)
    return send_file(buf, mimetype="image/jpeg")


@app.route("/results/<job_id>/update", methods=["PUT"])
def update_result(job_id):
    job = ocr_engine.get_job(job_id)
    if not job or job["status"] != "done":
        abort(404)

    data = request.get_json()
    page_idx = data.get("page", 0)
    elem_type = data.get("type")
    new_text = data.get("text", "")

    if page_idx >= len(job["results"]):
        abort(400)

    page_data = job["results"][page_idx]

    if elem_type == "paragraph":
        idx = data.get("index", 0)
        if idx < len(page_data.get("paragraphs", [])):
            page_data["paragraphs"][idx]["contents"] = new_text
    elif elem_type == "table":
        table_idx = data.get("table_index", 0)
        cell_idx = data.get("cell_index", 0)
        tables = page_data.get("tables", [])
        if table_idx < len(tables):
            cells = tables[table_idx].get("cells", [])
            if cell_idx < len(cells):
                cells[cell_idx]["contents"] = new_text

    return jsonify({"ok": True})


@app.route("/results/<job_id>/ai-correct", methods=["POST"])
def ai_correct(job_id):
    job = ocr_engine.get_job(job_id)
    if not job or job["status"] != "done":
        abort(404)

    data = request.get_json()
    page_idx = data.get("page", 0)

    if page_idx >= len(job["results"]):
        abort(400)

    try:
        corrections = ai_corrector.correct_and_update(job["results"][page_idx])
        return jsonify({"ok": True, "corrections": corrections, "page_data": job["results"][page_idx]})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/results/<job_id>/extract-structure", methods=["POST"])
def extract_structure(job_id):
    job = ocr_engine.get_job(job_id)
    if not job or job["status"] != "done":
        abort(404)

    data = request.get_json()
    page_idx = data.get("page", 0)

    if page_idx >= len(job["results"]):
        abort(400)

    try:
        # ページ画像をbase64エンコードしてVision APIに渡す
        import base64
        page_image_base64 = None
        if page_idx < len(job.get("images", [])):
            buf = ocr_engine.get_page_image_jpeg(job_id, page_idx)
            if buf:
                page_image_base64 = base64.b64encode(buf.read()).decode("ascii")

        structured = ai_corrector.extract_structured(job["results"][page_idx], page_image_base64)
        return jsonify({"ok": True, "structured": structured})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/results/<job_id>/ai-analyze", methods=["POST"])
def ai_analyze(job_id):
    """AI校正→構造化抽出を一括実行"""
    job = ocr_engine.get_job(job_id)
    if not job or job["status"] != "done":
        abort(404)

    data = request.get_json()
    page_idx = data.get("page", 0)

    if page_idx >= len(job["results"]):
        abort(400)

    try:
        import base64
        page_image_base64 = None
        if page_idx < len(job.get("images", [])):
            buf = ocr_engine.get_page_image_jpeg(job_id, page_idx)
            if buf:
                page_image_base64 = base64.b64encode(buf.read()).decode("ascii")

        corrections, structured = ai_corrector.analyze_page(
            job["results"][page_idx], page_image_base64
        )

        # 構造化データをジョブに保存（エクスポート用）
        if "structured" not in job:
            job["structured"] = {}
        job["structured"][str(page_idx)] = structured

        return jsonify({
            "ok": True,
            "corrections": corrections,
            "structured": structured,
            "page_data": job["results"][page_idx],
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/results/<job_id>/save-structured", methods=["PUT"])
def save_structured(job_id):
    """構造化ビューの編集内容を保存"""
    job = ocr_engine.get_job(job_id)
    if not job or job["status"] != "done":
        abort(404)

    data = request.get_json()
    page_idx = data.get("page", 0)
    structured = data.get("structured")

    if "structured" not in job:
        job["structured"] = {}
    job["structured"][str(page_idx)] = structured

    return jsonify({"ok": True})


@app.route("/export/<job_id>")
def export(job_id):
    job = ocr_engine.get_job(job_id)
    if not job or job["status"] != "done":
        abort(404)

    fmt = request.args.get("format", "json")
    results = job["results"]

    if fmt == "structured":
        structured = job.get("structured", {})
        buf = io.BytesIO(json.dumps(structured, ensure_ascii=False, indent=2).encode("utf-8"))
        buf.seek(0)
        return send_file(buf, mimetype="application/json",
                         as_attachment=True, download_name="structured.json")

    elif fmt == "json":
        return jsonify(results)

    elif fmt == "md":
        md = generate_markdown(results)
        buf = io.BytesIO(md.encode("utf-8"))
        buf.seek(0)
        return send_file(buf, mimetype="text/markdown",
                         as_attachment=True, download_name="ocr_result.md")

    elif fmt == "csv":
        csv_data = generate_csv(results)
        buf = io.BytesIO(csv_data.encode("utf-8-sig"))
        buf.seek(0)
        return send_file(buf, mimetype="text/csv",
                         as_attachment=True, download_name="ocr_tables.csv")

    elif fmt == "xlsx":
        buf = generate_xlsx(results)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="ocr_result.xlsx")

    return jsonify({"error": "Unknown format"}), 400


def generate_markdown(results):
    lines = []
    for page_idx, page in enumerate(results):
        lines.append(f"# Page {page_idx + 1}\n")

        elements = []
        for p in page.get("paragraphs", []):
            elements.append(("paragraph", p.get("order", 0), p))
        for t in page.get("tables", []):
            elements.append(("table", t.get("order", 999), t))
        elements.sort(key=lambda x: x[1])

        for etype, _, edata in elements:
            if etype == "paragraph":
                role = edata.get("role", "")
                text = edata.get("contents", "")
                if role == "section_heading":
                    lines.append(f"## {text}\n")
                else:
                    lines.append(f"{text}\n")
            else:
                n_row = edata.get("n_row", 0)
                n_col = edata.get("n_col", 0)
                if n_row == 0 or n_col == 0:
                    continue
                grid = [["" for _ in range(n_col)] for _ in range(n_row)]
                for cell in edata.get("cells", []):
                    r = (cell.get("row", 1)) - 1
                    c = (cell.get("col", 1)) - 1
                    grid[r][c] = cell.get("contents", "").replace("\n", " ")

                for ri, row in enumerate(grid):
                    lines.append("| " + " | ".join(row) + " |")
                    if ri == 0:
                        lines.append("| " + " | ".join(["---"] * n_col) + " |")
                lines.append("")

        lines.append("---\n")
    return "\n".join(lines)


def generate_csv(results):
    output = io.StringIO()
    writer = csv.writer(output)

    for page_idx, page in enumerate(results):
        for t_idx, table in enumerate(page.get("tables", [])):
            n_row = table.get("n_row", 0)
            n_col = table.get("n_col", 0)
            if n_row == 0 or n_col == 0:
                continue

            writer.writerow([f"=== Page {page_idx + 1}, Table {t_idx + 1} ==="])
            grid = [["" for _ in range(n_col)] for _ in range(n_row)]
            for cell in table.get("cells", []):
                r = (cell.get("row", 1)) - 1
                c = (cell.get("col", 1)) - 1
                grid[r][c] = cell.get("contents", "")
            for row in grid:
                writer.writerow(row)
            writer.writerow([])

    return output.getvalue()


def generate_xlsx(results):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR結果"

    row_offset = 1
    for page_idx, page in enumerate(results):
        ws.cell(row=row_offset, column=1, value=f"Page {page_idx + 1}")
        row_offset += 1

        # Paragraphs
        for p in page.get("paragraphs", []):
            text = p.get("contents", "")
            role = p.get("role", "")
            ws.cell(row=row_offset, column=1, value=f"[{role}]" if role else "")
            ws.cell(row=row_offset, column=2, value=text)
            row_offset += 1

        row_offset += 1

        # Tables
        for t_idx, table in enumerate(page.get("tables", [])):
            n_row = table.get("n_row", 0)
            n_col = table.get("n_col", 0)
            if n_row == 0 or n_col == 0:
                continue

            ws.cell(row=row_offset, column=1, value=f"Table {t_idx + 1}")
            row_offset += 1

            grid = [["" for _ in range(n_col)] for _ in range(n_row)]
            for cell in table.get("cells", []):
                r = (cell.get("row", 1)) - 1
                c = (cell.get("col", 1)) - 1
                grid[r][c] = cell.get("contents", "")

            for grid_row in grid:
                for ci, val in enumerate(grid_row):
                    ws.cell(row=row_offset, column=ci + 1, value=val)
                row_offset += 1

            row_offset += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============ 相談シートモード ============

@app.route("/upload-consultation", methods=["POST"])
def upload_consultation():
    files = request.files.getlist("files")
    if not files:
        return redirect(url_for("index"))

    file_entries = []
    batch_dir = os.path.join(app.config["UPLOAD_FOLDER"], f"batch_{int(time.time())}")
    os.makedirs(batch_dir, exist_ok=True)

    for f in files:
        if not f.filename:
            continue
        ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""

        if ext == "zip":
            # Extract ZIP and find PDFs inside
            zip_path = os.path.join(batch_dir, f"temp_{int(time.time())}.zip")
            f.save(zip_path)
            with zipfile.ZipFile(zip_path, "r") as zf:
                for name in zf.namelist():
                    if name.lower().endswith(".pdf") and not name.startswith("__MACOSX"):
                        basename = Path(name).name
                        extract_path = os.path.join(batch_dir, basename)
                        with zf.open(name) as src, open(extract_path, "wb") as dst:
                            dst.write(src.read())
                        file_entries.append({"name": basename, "path": extract_path})
            os.remove(zip_path)

        elif ext == "pdf":
            safe_name = f"upload_{int(time.time())}_{len(file_entries)}.pdf"
            file_path = os.path.join(batch_dir, safe_name)
            f.save(file_path)
            file_entries.append({"name": f.filename, "path": file_path})

    if not file_entries:
        return redirect(url_for("index"))

    batch_id = ocr_engine.create_batch(file_entries)
    return redirect(url_for("consultation_processing", batch_id=batch_id))


@app.route("/consultation/processing/<batch_id>")
def consultation_processing(batch_id):
    batch = ocr_engine.get_batch(batch_id)
    if not batch:
        abort(404)
    # Check if all done
    all_done = all(
        (ocr_engine.get_job(jid) or {}).get("status") == "done"
        for jid in batch["job_ids"]
    )
    if all_done:
        return redirect(url_for("consultation_results", batch_id=batch_id))
    return render_template("consultation_processing.html", batch_id=batch_id)


@app.route("/consultation/status/<batch_id>")
def consultation_status(batch_id):
    batch = ocr_engine.get_batch(batch_id)
    if not batch:
        return jsonify({"error": "Batch not found"}), 404

    items = []
    for jid in batch["job_ids"]:
        job = ocr_engine.get_job(jid) or {}
        items.append({
            "job_id": jid,
            "name": job.get("source_name", ""),
            "status": job.get("status", "pending"),
            "current_page": job.get("current_page", 0),
            "page_count": job.get("page_count", 0),
            "error": job.get("error"),
        })

    all_done = all(it["status"] in ("done", "error") for it in items)
    return jsonify({"items": items, "all_done": all_done})


@app.route("/consultation/results/<batch_id>")
def consultation_results(batch_id):
    batch = ocr_engine.get_batch(batch_id)
    if not batch:
        abort(404)

    # Collect job data
    jobs_data = []
    for jid in batch["job_ids"]:
        job = ocr_engine.get_job(jid)
        if not job or job["status"] != "done":
            continue
        jobs_data.append({
            "job_id": jid,
            "name": job.get("source_name", jid),
            "page_count": len(job["results"]),
            "ocr_data": job["results"],
        })

    return render_template(
        "consultation_results.html",
        batch_id=batch_id,
        jobs_data=jobs_data,
    )


@app.route("/consultation/<batch_id>/ai-analyze-job", methods=["POST"])
def consultation_ai_analyze(batch_id):
    """相談シート専用 AI解析（1ジョブ分）"""
    batch = ocr_engine.get_batch(batch_id)
    if not batch:
        abort(404)

    data = request.get_json()
    job_id = data.get("job_id")
    if job_id not in batch["job_ids"]:
        abort(400)

    job = ocr_engine.get_job(job_id)
    if not job or job["status"] != "done":
        abort(400)

    try:
        import base64
        # Use first page (consultation sheets are typically 1-2 pages)
        page_idx = 0
        page_image_base64 = None
        if page_idx < len(job.get("images", [])):
            buf = ocr_engine.get_page_image_jpeg(job_id, page_idx)
            if buf:
                page_image_base64 = base64.b64encode(buf.read()).decode("ascii")

        structured = ai_corrector.extract_consultation_structured(
            job["results"][page_idx], page_image_base64
        )

        # Save structured data to batch
        with ocr_engine.batches_lock:
            ocr_engine.batches[batch_id]["structured"][job_id] = structured

        return jsonify({"ok": True, "structured": structured})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/consultation/<batch_id>/save-structured", methods=["PUT"])
def consultation_save_structured(batch_id):
    """構造化ビューの編集内容を保存"""
    batch = ocr_engine.get_batch(batch_id)
    if not batch:
        abort(404)

    data = request.get_json()
    job_id = data.get("job_id")
    structured = data.get("structured")

    with ocr_engine.batches_lock:
        ocr_engine.batches[batch_id]["structured"][job_id] = structured

    return jsonify({"ok": True})


@app.route("/consultation/<batch_id>/export-csv")
def consultation_export_csv(batch_id):
    """DentNet CSV一括エクスポート"""
    batch = ocr_engine.get_batch(batch_id)
    if not batch:
        abort(404)

    import consultation_csv
    rows = []
    for jid in batch["job_ids"]:
        struct = batch["structured"].get(jid)
        if struct:
            rows.append(consultation_csv.structured_to_dentnet_row(struct))

    csv_bytes = consultation_csv.rows_to_csv_bytes(rows)
    buf = io.BytesIO(csv_bytes)
    buf.seek(0)
    return send_file(buf, mimetype="text/csv",
                     as_attachment=True,
                     download_name="dentnet_patient_data.csv")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5002))
    app.run(host="0.0.0.0", port=port, debug=False)
